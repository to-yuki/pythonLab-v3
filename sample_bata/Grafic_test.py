# -*- coding: UTF-8 -*-
import openpyxl
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.font_manager as mf
import pprint

#日本語をグラフに付けるため、使えるフォントを確認する（参考）
fonts = mf.findSystemFonts()
# pprint.pprint([[str(font), mf.FontProperties(fname=font).get_name()] for font in fonts[:]])　#pprint はMapをきれいに出力できる

# pltのフォントの設定（参考）
font = {'family' : 'Century Gothic',
       'weight' : 'bold',
        'size'   : 10}
matplotlib.rc('font', **font)

# Excel ファイルをロード  
book = openpyxl.load_workbook(filename="test.xlsx",data_only=True)
# Excel シート名のリスト取得
sheetnames = book.get_sheet_names()

# 1番目のシートにアクセス
sheet = book.get_sheet_by_name(sheetnames[0])

# 各シートのデータ数表示
for sheetname in sheetnames:
    sheet = book.get_sheet_by_name(sheetname)
    print(sheetname + u" のデータ数: "),
    print(u"行:" + str(sheet.max_row)),
    print(u"列:" + str(sheet.max_column))

# 1番目のシートにアクセス
sheet = book.get_sheet_by_name(sheetnames[0])
# A1 のセルデータを取得
data = sheet["A1"].value
print(u"A1: {}".format(data))

# シート状のすべてのデータを表示
data_list = []
for r in range(1,sheet.max_row+1):
    print()
    for c in range(1,sheet.max_column+1):
        data = sheet.cell(row=r,column=c).value
        if data != None:
            # cell の列と行の表示(列はchrで数値を文字に変換)
            print(chr(c+64) + str(r) + ":", end = " ")
            print(data, end = "\n")

# グラフ作成
labels = []
lang_val = []
math_val = []
lang = sheet.cell(row=1,column=2).value
math = sheet.cell(row=1,column=3).value
x = 1
for r in range(2, sheet.max_row):
    labels.append(sheet.cell(row=r,column=1).value)
    lang_val.append(sheet.cell(row=r,column=2).value)
    math_val.append(sheet.cell(row=r,column=3).value)
    x += 1

print()
print("国語： {}".format(lang_val))
print("数学： {}".format(math_val))

# 積み上げ棒グラフ　日本語の出力は難しいので、半角で出力
lang_test = plt.bar(range(1, x), lang_val, color = "b")
math_test = plt.bar(range(1, x), math_val, color = "c", bottom = lang_val)
plt.xticks(range(1, x), labels)
plt.legend((lang_test, math_test), (lang, math))
plt.show()

# 散布図
plt.scatter(lang_val, math_val)
plt.xlabel("Japanese")
plt.ylabel("Math")
plt.show()

# 箱ひげ図
point = (lang_val, math_val)
fig = plt.figure()
ax = fig.add_subplot(111)
bp = ax.boxplot(point)
ax.set_xticklabels(["Japanese", "Math"])
plt.show()