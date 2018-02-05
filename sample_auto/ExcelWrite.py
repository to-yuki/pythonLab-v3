# -*- coding: UTF-8 -*-
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side

# 新規のワークブックの作成
book = openpyxl.Workbook()
# 新規作成された最初のシートを取得
sheet = book.active
# シート名を変更(Sheet1 -> シート1)
sheet.title = u"基本操作サンプル"
# A1 セルに値を設定
sheet["B2"] = u"ハロー"

# A2 のセル以下 A100 まで繰返しデータの設定
for i in range(2,101):
    sheet.cell(row=i,column=1).value = u"Python ワールド"

# 新規シートの作成
newSheet = book.create_sheet(index=1,title=u"フォントサンプル")
newSheet["B2"] = u"Python入門"

# セルのプロパティオブジェクトの作成
# フォントオブジェクト
customFont = Font(name=u"メイリオ",bold=True,italic=True,size=24,color="00F00000")
# フィルオブジェクト
customFill = PatternFill(patternType="solid", start_color="30F0F0F0", end_color="BBBB0000")
# 罫線のオブジェクト
customborder = Border(outline=True,
                        left=Side(style="thin", color="FF000000"),
                        right=Side(style="thin", color="FF000000"),
                        top=Side(style="thin", color="FF000000"),
                        bottom=Side(style="thin", color="FF000000"))
# B2 セルに各オブジェクトを設定して調整
newSheet[u"B2"].font = customFont
newSheet[u"B2"].fill = customFill
newSheet[u"B2"].border = customborder

# セルの幅調整
newSheet.row_dimensions[2].height=60
newSheet.column_dimensions[u"B"].width=30

# ワークブックの保存
book.save("newBook.xlsx")
