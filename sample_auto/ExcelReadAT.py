# -*- coding: UTF-8 -*-
import openpyxl

# Excel ファイルをロード  
book = openpyxl.load_workbook(filename="test.xlsx",data_only=True)
#book = openpyxl.load_workbook(filename="test.xlsx",data_only=False)

# Excel シート名のリスト取得
sheetnames = book.get_sheet_names()
# 各シートのデータ数表示
for sheetname in sheetnames:
    sheet = book.get_sheet_by_name(sheetname)
    print(sheetname + u" のデータ数: "),
    print(u"行:" + str(sheet.max_row)),
    print(u"列:" + str(sheet.max_column))

# 1番目のシートにアクセス
sheet = book.get_sheet_by_name(sheetnames[0])
# A1 のセルデータを取得
data = sheet["A1"].value
print("A1: " ),
print(data)

# シート状のすべてのデータを表示
for r in range(1,sheet.max_row+1):
    for c in range(1,sheet.max_column+1):
        data = sheet.cell(row=r,column=c).value
        if data != None:
            # cell の列と行の表示(列はchrで数値を文字に変換)
            print(chr(c+64) + str(r) + ":"),
            print(data)
