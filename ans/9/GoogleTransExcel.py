# -*- coding: UTF-8 -*-
import openpyxl
import requests
import re
from googletrans import Translator

# Google Translator インスタンス化
translator = Translator() 
# Excel ファイルをロード  
book = openpyxl.load_workbook(filename='GoogleTransExcel.xlsx',data_only=True)
# Excel シート名のリスト取得
sheetnames = book.sheetnames
# 1番目のシートにアクセス
sheet = book[sheetnames[0]]

list = []

for r in range(1,sheet.max_row+1):
    data = sheet.cell(row=r,column=1).value
    list.append(data)

# 行カウンタのリセット
r = 0
# 翻訳エラー時のリトライ回数
retrycount = 3

for origin in list:
    if (origin == None or r == 0 ):
        r+=1
        continue
    # 翻訳失敗時は retrycount 回リトライする
    while(retrycount != 0):
        try:
            # resopnse 内にある JavaScript コードから翻訳された文字列を取得
            transValue = translator.translate(text=origin, dest='ja')
            #print(transValue)
            # ExcelSheetに書き込み(list->0,excel->1)
            # 翻訳テキストの書き込み列は column で指定
            sheet.cell(row=r+1,column=2).value = transValue.text
            print
            print(str(r)+'行目 [translate.google.com] で翻訳') 
            print('\t英語:'),
            print(origin)
            print('\t日本語:'),
            print(transValue.text)
            retrycount = 3
            break
        except:
            print(str(r)+'行目翻訳リトライ。')
            retrycount -= 1
    else:
        print(str(r)+'行目:翻訳エラーのリトライ回数をオーバーしたので、スキップします。')
    r+=1
# ワークブックの保存
try:
    book.save('GoogleTransExcel-OK.xlsx')
except:
    print('Excelファイル保存エラーです。')